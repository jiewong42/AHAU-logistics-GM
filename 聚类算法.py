import pandas as pd
import numpy as np
import openpyxl
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties, fontManager
from matplotlib.ticker import MultipleLocator

# 查找并选择正确的中文字体
font_path = None
for font in fontManager.ttflist:
    if 'SimSun' in font.name:
        font_path = font.fname
        break

# 设置中文字体
font = FontProperties(fname=font_path, size=12)

# 从Excel文件中读取数据
data = pd.read_excel("mdl4.xlsx")

# 提取经度和纬度列数据
coordinates = data.iloc[:, 1:].values

# 特征缩放
scaler = StandardScaler()
scaled_coordinates = scaler.fit_transform(coordinates)

# 设置聚类数目
k = 14

# 运行K-means算法
kmeans = KMeans(n_clusters=k, random_state=42)
kmeans.fit(scaled_coordinates)

# 反向转换聚类中心坐标到原始范围
cluster_centers = scaler.inverse_transform(kmeans.cluster_centers_)

# 输出每个聚类中心坐标
for i, center in enumerate(cluster_centers):
    print("聚类中心", i+1, "经度:", center[0], "纬度:", center[1])

# 创建一个空的字典，用于存储每个聚类中心的点
cluster_points = {i: [] for i in range(k)}

# 将每个数据点分配到相应的聚类中心
labels = kmeans.labels_
for i, label in enumerate(labels):
    cluster_points[label].append(coordinates[i])

# 输出每个聚类中心的所有点到Excel文件
wb = openpyxl.Workbook()
for i in range(k):
    cluster_data = np.array(cluster_points[i])
    sheet = wb.create_sheet(title='Cluster {}'.format(i+1))
    for j, coord in enumerate(cluster_data):
        sheet.cell(row=j+1, column=1).value = coord[0]
        sheet.cell(row=j+1, column=2).value = coord[1]

# 保存Excel文件
wb.save('cluster_points.xlsx')

# 设置图形的大小和分辨率
fig, ax = plt.subplots(figsize=(12, 9), dpi=300)

# 绘制数据点散点图
scatter = ax.scatter(
    coordinates[:, 0],
    coordinates[:, 1],
    c=labels,
    cmap='viridis',
    alpha=0.8,
    s=50,
    linewidths=1
)

# 绘制聚类中心点散点图
ax.scatter(
    cluster_centers[:, 0],
    cluster_centers[:, 1],
    marker='x',
    c='red',
    s=100
)

# 自定义图表
ax.set_xlabel('经度', fontproperties=font)
ax.set_ylabel('纬度', fontproperties=font)
ax.set_title('麦当劳选址聚类结果', fontproperties=font)

# 设置刻度线密度
ax.xaxis.set_major_locator(MultipleLocator(0.2))
ax.yaxis.set_major_locator(MultipleLocator(0.2))

# 添加颜色条图例
cbar = plt.colorbar(scatter)
cbar.set_label('Cluster')

# 调整图表布局并显示图表
plt.tight_layout()
plt.show()
